import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.shortcuts import render
from django.http import HttpResponse
from django.core.files.base import ContentFile
from .forms import UploadFileForm
import io

def handle_uploaded_file(file):
    pass

def process_excel(file, action):
    if action == 'action1':  # ODD semester
        return process_odd_semester(file)
    elif action == 'action2':  # EVEN semester
        return process_even_semester(file)
    else:
        raise ValueError("Invalid action selected")

def process_odd_semester(file):
    df = pd.read_excel(file, usecols="C:E,P,Q", skiprows=3, names=["SEM", "Course", "Branch", "Gender", "Category"])
    semesters = [1, 3, 5, 7]
    return process_semesters(df, semesters)

def process_even_semester(file):
    df = pd.read_excel(file, usecols="C:E,P,Q", skiprows=3, names=["SEM", "Course", "Branch", "Gender", "Category"])
    semesters = [2, 4, 6, 8]
    return process_semesters(df, semesters)

def process_semesters(df, semesters):
    results = []
    for sem in semesters:
        df_filtered_A = df[(df['SEM'] == sem) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'AGRICULTURE ENGINEERING')]
        df_filtered_C = df[(df['SEM'] == sem) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'CIVIL')]
        df_filtered_E = df[(df['SEM'] == sem) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'ELECTRICAL')]
        df_filtered_F = df[(df['SEM'] == sem) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'FOOTWEAR TECHNOLOGY')]
        df_filtered_M = df[(df['SEM'] == sem) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'MECHANICAL')]

        m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
        m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
        m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
        m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
        m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
        Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

        f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
        f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
        f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
        f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
        f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
        Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

        total_A = m_sum_A + f_sum_A
        total_C = m_sum_C + f_sum_C
        total_E = m_sum_E + f_sum_E
        total_F = m_sum_F + f_sum_F
        total_M = m_sum_M + f_sum_M
        Total_NO_All = total_A + total_C + total_E + total_F + total_M

        gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
        gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
        gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
        gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
        gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
        gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

        bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
        bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
        bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
        bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
        bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
        bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

        sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
        sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
        sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
        sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
        sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
        sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

        st_count_A = df_filtered_A['Category'].str.count('ST').sum()
        st_count_C = df_filtered_C['Category'].str.count('ST').sum()
        st_count_E = df_filtered_E['Category'].str.count('ST').sum()
        st_count_F = df_filtered_F['Category'].str.count('ST').sum()
        st_count_M = df_filtered_M['Category'].str.count('ST').sum()
        st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

        CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
        CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
        CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
        CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
        CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
        CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

        data = {
            'SEM': [sem, sem, sem, sem, sem, ' ', ''],
            'Course': ['B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', ' ', ''],
            'Branch': ['AGRICULTURE ENGINEERING', 'CIVIL', 'ELECTRICAL', 'FOOTWEAR TECHNOLOGY', 'MECHANICAL', 'TOTAL', ''],
            'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M, Total_NO_m, ''],
            'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M, Total_NO_f, ''],
            'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M, Total_NO_All, ''],
            'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M, gn_total, ''],
            'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M, bc_total, ''],
            'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M, sc_total, ''],
            'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M, st_total, ''],
            'TOTAL(C)': [CA_total_A, CA_total_C, CA_total_E, CA_total_F, CA_total_M, CA_total_All, '']
        }
        results.append(pd.DataFrame(data))

    df_combined = pd.concat(results, ignore_index=True)

    wb = load_workbook(filename=file)
    sheet_name = 'Sheet3'
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    ws = wb[sheet_name]
    for row in dataframe_to_rows(df_combined, index=False, header=True):
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            file_contents = uploaded_file.read()
            action = request.POST.get('action')
            
            # Process the uploaded file based on the selected action
            processed_data = process_excel(io.BytesIO(file_contents), action)
            
            # Prepare response for downloading the processed file
            response = HttpResponse(processed_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=New.xlsx'
            return response
    else:
        form = UploadFileForm()
    return render(request, 'upload.html', {'form': form})