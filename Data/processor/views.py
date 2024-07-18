import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.shortcuts import render
from django.http import HttpResponse
from django.core.files.base import ContentFile
from .forms import UploadFileForm
import io

def handle_uploaded_file(file):
    pass

def process_excel(file):
    df = pd.read_excel(file, usecols="C:E,P,Q", skiprows=3, names=["SEM", "Course", "Branch", "Gender", "Category"])

    # Your existing processing code...
    df_filtered_A = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'AGRICULTURAL ENGINEERING')]
    df_filtered_C = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'CIVIL') ]
    df_filtered_E = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'ELECTRICAL')]
    df_filtered_F = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'FOOTWEAR TECHNOLOGY')]
    df_filtered_M = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'MECHANICAL')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    data_1 = {
        'SEM': [2, 2, 2, 2, 2, ' ', ''],
        'Course': ['B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', ' ', ''],
        'Branch': ['AGRICULTURAL ENGINEERING', 'CIVIL', 'ELECTRICAL', 'FOOTWEAR TECHNOLOGY', 'MECHANICAL', 'TOTAL', ''],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M, Total_NO_m, ''],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M, Total_NO_f, ''],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M, Total_NO_All, ''],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M, gn_total, ''],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M, bc_total, ''],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M, sc_total, ''],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M, st_total, ''],
        'TOTAL(C)': [CA_total_A, CA_total_C, CA_total_E, CA_total_F, CA_total_M, CA_total_All, '']
    }

    df_data_1 = pd.DataFrame(data_1)

    df_filtered_A = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'AGRICULTURAL ENGINEERING')]
    df_filtered_C = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'CIVIL') ]
    df_filtered_E = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'ELECTRICAL')]
    df_filtered_F = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'FOOTWEAR TECHNOLOGY')]
    df_filtered_M = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'MECHANICAL')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    data_2 = {
        'SEM': [4, 4, 4, 4, 4, ' ', ''],
        'Course': ['B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', ' ', ''],
        'Branch': ['AGRICULTURAL ENGINEERING', 'CIVIL', 'ELECTRICAL', 'FOOTWEAR TECHNOLOGY', 'MECHANICAL', 'TOTAL', ''],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M, Total_NO_m, ''],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M, Total_NO_f, ''],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M, Total_NO_All, ''],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M, gn_total, ''],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M, bc_total, ''],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M, sc_total, ''],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M, st_total, ''],
        'TOTAL(C)': [CA_total_A, CA_total_C, CA_total_E, CA_total_F, CA_total_M, CA_total_All, '']
    }

    df_data_2 = pd.DataFrame(data_2)

    
    df_combined = pd.concat([df_data_1, df_data_2], ignore_index=True)

    
    wb = load_workbook(filename=file)
    
    
    sheet_name = 'Sheet3'
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    
    # Select the sheet for writing
    ws = wb[sheet_name]
    
    # Append data to the sheet
    for row in dataframe_to_rows(df_combined, index=False, header=True):
        ws.append(row)
    
    # Save changes to the workbook
    wb.save(file)

    # Return the file path or bytesIO object for downloading or further processing
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            file_contents = uploaded_file.read()
            
            # Process the uploaded file
            processed_data = process_excel(io.BytesIO(file_contents))
            
            # Prepare response for downloading the processed file
            response = HttpResponse(processed_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; New.xlsx'
            return response
    else:
        form = UploadFileForm()
    return render(request, 'upload.html', {'form': form})