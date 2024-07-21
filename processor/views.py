import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.shortcuts import render
from django.http import HttpResponse
from django.core.files.base import ContentFile
from .forms import UploadFileForm
import io



def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            file_contents = uploaded_file.read()
            action = form.cleaned_data['action']

            
            if action == 'action1':
                # Logic for ODD action
                processed_data = process_excel_odd(io.BytesIO(file_contents))
            elif action == 'action2':
                # Logic for EVEN action
                processed_data = process_excel_even(io.BytesIO(file_contents))
            else:
                # Handle unexpected action
                return HttpResponse("Invalid action selected", status=400)

            # Prepare response for downloading the processed file
            response = HttpResponse(processed_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=New.xlsx'
            return response
    else:
        form = UploadFileForm()
    return render(request, 'upload.html', {'form': form})

def process_excel_odd(file):
    df = pd.read_excel(file, usecols="C:E,P,Q", skiprows=3, names=["SEM", "Course", "Branch", "Gender", "Category"])


    df_filtered_A = df[(df['SEM'] == 1) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'AGRICULTURE ENGINEERING')]
    df_filtered_C = df[(df['SEM'] == 1) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'CIVIL') ]
    df_filtered_E = df[(df['SEM'] == 1) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'ELECTRICAL')]
    df_filtered_F = df[(df['SEM'] == 1) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'FOOTWEAR TECHNOLOGY')]
    df_filtered_M = df[(df['SEM'] == 1) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'MECHANICAL')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    d1_Total_NO_m = Total_NO_m
    d1_Total_NO_f = Total_NO_f
    d1_Total_NO_All = Total_NO_All
    d1_gn_total = gn_total
    d1_bc_total = bc_total
    d1_sc_total = sc_total
    d1_st_total = st_total
    d1_CA_total_All = CA_total_All

    data_1 = {
        'SEM': [1, 1, 1, 1, 1 ,' ',''],
        'Course': ['B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', 'B.Tech',' ',''],
        'Branch': ['AGRICULTURE ENGINEERING', 'CIVIL', 'ELECTRICAL', 'FOOTWEAR TECHNOLOGY', 'MECHANICAL','TOTAL',''],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,''],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,''],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,''],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total,''],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total,''],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total,''],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total,''],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All,'']
    }


    df_data_1 = pd.DataFrame(data_1)



    df_filtered_A = df[(df['SEM'] == 3) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'AGRICULTURE ENGINEERING')]
    df_filtered_C = df[(df['SEM'] == 3) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'CIVIL') ]
    df_filtered_E = df[(df['SEM'] == 3) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'ELECTRICAL')]
    df_filtered_F = df[(df['SEM'] == 3) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'FOOTWEAR TECHNOLOGY')]
    df_filtered_M = df[(df['SEM'] == 3) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'MECHANICAL')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    d2_Total_NO_m = Total_NO_m
    d2_Total_NO_f = Total_NO_f
    d2_Total_NO_All = Total_NO_All
    d2_gn_total = gn_total
    d2_bc_total = bc_total
    d2_sc_total = sc_total
    d2_st_total = st_total
    d2_CA_total_All = CA_total_All

    data_2 = {
        'SEM': [3, 3, 3, 3, 3 ,' ',''],
        'Course': ['B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', 'B.Tech',' ',''],
        'Branch': ['AGRICULTURE ENGINEERING', 'CIVIL', 'ELECTRICAL', 'FOOTWEAR TECHNOLOGY', 'MECHANICAL','TOTAL',''],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,''],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,''],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,''],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total,''],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total,''],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total,''],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total,''],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All,'']
    }


    df_data_2 = pd.DataFrame(data_2)

    df_filtered_A = df[(df['SEM'] == 5) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'AGRICULTURE ENGINEERING')]
    df_filtered_C = df[(df['SEM'] == 5) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'CIVIL') ]
    df_filtered_E = df[(df['SEM'] == 5) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'ELECTRICAL')]
    df_filtered_F = df[(df['SEM'] == 5) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'FOOTWEAR TECHNOLOGY')]
    df_filtered_M = df[(df['SEM'] == 5) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'MECHANICAL')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    d3_Total_NO_m = Total_NO_m
    d3_Total_NO_f = Total_NO_f
    d3_Total_NO_All = Total_NO_All
    d3_gn_total = gn_total
    d3_bc_total = bc_total
    d3_sc_total = sc_total
    d3_st_total = st_total
    d3_CA_total_All = CA_total_All

    data_3 = {
        'SEM': [5, 5, 5, 5, 5 ,' ',''],
        'Course': ['B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', 'B.Tech',' ',''],
        'Branch': ['AGRICULTURE ENGINEERING', 'CIVIL', 'ELECTRICAL', 'FOOTWEAR TECHNOLOGY', 'MECHANICAL','TOTAL',''],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,''],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,''],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,''],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total,''],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total,''],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total,''],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total,''],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All,'']
    }

    df_data_3 = pd.DataFrame(data_3)

    df_filtered_A = df[(df['SEM'] == 7) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'AGRICULTURE ENGINEERING')]
    df_filtered_C = df[(df['SEM'] == 7) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'CIVIL') ]
    df_filtered_E = df[(df['SEM'] == 7) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'ELECTRICAL')]
    df_filtered_F = df[(df['SEM'] == 7) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'FOOTWEAR TECHNOLOGY')]
    df_filtered_M = df[(df['SEM'] == 7) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'MECHANICAL')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M


    d4_Total_NO_m = Total_NO_m
    d4_Total_NO_f = Total_NO_f
    d4_Total_NO_All = Total_NO_All
    d4_gn_total = gn_total
    d4_bc_total = bc_total
    d4_sc_total = sc_total
    d4_st_total = st_total
    d4_CA_total_All = CA_total_All

    Grand_Total_CA_total_All = GT_CA_total_All = d1_CA_total_All + d2_CA_total_All + d3_CA_total_All + d4_CA_total_All
    Grand_Total_st_total  = GT_st_total = d1_st_total + d2_st_total + d3_st_total + d4_st_total
    Grand_Total_sc_total  = GT_sc_total = d1_sc_total + d2_sc_total + d3_sc_total + d4_sc_total
    Grand_Total_bc_total  = GT_bc_total = d1_bc_total + d2_bc_total + d3_bc_total + d4_bc_total
    Grand_Total_gn_total  = GT_gn_total = d1_gn_total + d2_gn_total + d3_gn_total + d4_gn_total
    Grand_Total_Total_NO_All = GT_Total_NO_All = d1_Total_NO_All + d2_Total_NO_All + d3_Total_NO_All + d4_Total_NO_All
    Grand_Total_Total_NO_f = GT_Total_NO_f = d1_Total_NO_f + d2_Total_NO_f + d3_Total_NO_f + d4_Total_NO_f
    Grand_Total_Total_NO_m = GT_Total_NO_m = d1_Total_NO_m + d2_Total_NO_m + d3_Total_NO_m + d4_Total_NO_m

    data_4 = {
        'SEM': [7, 7, 7, 7, 7 ,' ',''],
        'Course': ['B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', 'B.Tech',' ',''],
        'Branch': ['AGRICULTURE ENGINEERING', 'CIVIL', 'ELECTRICAL', 'FOOTWEAR TECHNOLOGY', 'MECHANICAL','TOTAL','GRAND TOTAL'],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,GT_Total_NO_m],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,GT_Total_NO_f],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,GT_Total_NO_All],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total ,GT_gn_total],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total , GT_bc_total],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total ,GT_sc_total],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total , GT_st_total],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All , GT_CA_total_All],

    }


    df_data_4 = pd.DataFrame(data_4)

   

    df_filtered_A = df[(df['SEM'] == 1) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'AI & ROBOTICS')]
    df_filtered_C = df[(df['SEM'] == 1) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'AUTOMOBILE') ]
    df_filtered_E = df[(df['SEM'] == 1) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'DIGITAL MANUFACTURING')]
    df_filtered_F = df[(df['SEM'] == 1) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'RENEWABLE ENERGY')]
    df_filtered_M = df[(df['SEM'] == 1) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'WATER, SANITATION AND WASTE MANAGEMENT')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    d1_Total_NO_m = Total_NO_m
    d1_Total_NO_f = Total_NO_f
    d1_Total_NO_All = Total_NO_All
    d1_gn_total = gn_total
    d1_bc_total = bc_total
    d1_sc_total = sc_total
    d1_st_total = st_total
    d1_CA_total_All = CA_total_All

    data_5 = {
        'SEM': [1, 1, 1, 1, 1 ,' ',''],
        'Course': ['B.Voc', 'B.Voc', 'B.Voc', 'B.Voc', 'B.Voc',' ',''],
        'Branch': ['AI & ROBxOTICS', 'AUTOMOBILE', 'DIGITAL MANUFACTURING', 'RENEWABLE ENERGY', 'WATER, SANITATION AND WASTE MANAGEMENT','TOTAL',''],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,''],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,''],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,''],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total,''],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total,''],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total,''],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total,''],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All,'']
    }

    df_data_5 = pd.DataFrame(data_5)


    df_filtered_A = df[(df['SEM'] == 3) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'AI & ROBOTICS')]
    df_filtered_C = df[(df['SEM'] == 3) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'AUTOMOBILE') ]
    df_filtered_E = df[(df['SEM'] == 3) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'DIGITAL MANUFACTURING')]
    df_filtered_F = df[(df['SEM'] == 3) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'RENEWABLE ENERGY')]
    df_filtered_M = df[(df['SEM'] == 3) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'WATER, SANITATION AND WASTE MANAGEMENT')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    d2_Total_NO_m = Total_NO_m
    d2_Total_NO_f = Total_NO_f
    d2_Total_NO_All = Total_NO_All
    d2_gn_total = gn_total
    d2_bc_total = bc_total
    d2_sc_total = sc_total
    d2_st_total = st_total
    d2_CA_total_All = CA_total_All

    data_6 = {
        'SEM': [3, 3, 3, 3, 3 ,' ',''],
        'Course': ['B.Voc', 'B.Voc', 'B.Voc', 'B.Voc', 'B.Voc',' ',''],
        'Branch': ['AI & ROBxOTICS', 'AUTOMOBILE', 'DIGITAL MANUFACTURING', 'RENEWABLE ENERGY', 'WATER, SANITATION AND WASTE MANAGEMENT','TOTAL',''],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,''],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,''],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,''],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total,''],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total,''],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total,''],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total,''],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All,'']
    }
    df_data_6 = pd.DataFrame(data_6)

    df_filtered_A = df[(df['SEM'] == 5) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'AI & ROBOTICS')]
    df_filtered_C = df[(df['SEM'] == 5) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'AUTOMOBILE') ]
    df_filtered_E = df[(df['SEM'] == 5) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'DIGITAL MANUFACTURING')]
    df_filtered_F = df[(df['SEM'] == 5) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'RENEWABLE ENERGY')]
    df_filtered_M = df[(df['SEM'] == 5) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'WATER, SANITATION AND WASTE MANAGEMENT')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    d3_Total_NO_m = Total_NO_m
    d3_Total_NO_f = Total_NO_f
    d3_Total_NO_All = Total_NO_All
    d3_gn_total = gn_total
    d3_bc_total = bc_total
    d3_sc_total = sc_total
    d3_st_total = st_total
    d3_CA_total_All = CA_total_All

    GT_CA_total_All = d1_CA_total_All + d2_CA_total_All + d3_CA_total_All
    GT_st_total = d1_st_total + d2_st_total + d3_st_total
    GT_sc_total = d1_sc_total + d2_sc_total + d3_sc_total
    GT_bc_total = d1_bc_total + d2_bc_total + d3_bc_total
    GT_gn_total = d1_gn_total + d2_gn_total + d3_gn_total
    GT_Total_NO_All = d1_Total_NO_All + d2_Total_NO_All + d3_Total_NO_All
    GT_Total_NO_f = d1_Total_NO_f + d2_Total_NO_f + d3_Total_NO_f
    GT_Total_NO_m = d1_Total_NO_m + d2_Total_NO_m + d3_Total_NO_m

    Grand_Total_CA_total_All = GT_CA_total_All + Grand_Total_CA_total_All
    Grand_Total_st_total  = GT_st_total + Grand_Total_st_total
    Grand_Total_sc_total  = GT_sc_total + Grand_Total_sc_total
    Grand_Total_bc_total  = GT_bc_total + Grand_Total_bc_total
    Grand_Total_gn_total  = GT_gn_total + Grand_Total_gn_total
    Grand_Total_Total_NO_All = GT_Total_NO_All + Grand_Total_Total_NO_All
    Grand_Total_Total_NO_f = GT_Total_NO_f + Grand_Total_Total_NO_f
    Grand_Total_Total_NO_m = GT_Total_NO_m + Grand_Total_Total_NO_m

    data_7 = {
        'SEM': [5, 5, 5, 5, 5 ,' ',''],
        'Course': ['B.Voc', 'B.Voc', 'B.Voc', 'B.Voc', 'B.Voc',' ',''],
        'Branch': ['AI & ROBxOTICS', 'AUTOMOBILE', 'DIGITAL MANUFACTURING', 'RENEWABLE ENERGY', 'WATER, SANITATION AND WASTE MANAGEMENT','TOTAL','GRAND TOTAL'],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,GT_Total_NO_m],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,GT_Total_NO_f],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,GT_Total_NO_All],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total ,GT_gn_total],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total , GT_bc_total],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total ,GT_sc_total],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total , GT_st_total],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All , GT_CA_total_All]
    }
    df_data_7 = pd.DataFrame(data_7)





    #----------------------------------------------------------------
    # B.Tech P/T
    #----------------------------------------------------------------




    df_filtered_E = df[(df['SEM'] == 1) & (df['Course'].isin(['B.Tech PT'])) & (df['Branch'] == 'ELECTRICAL')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E
    Total_NO_All =  total_E

    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_total = gn_count_E

    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_total =  bc_count_E

    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_total = sc_count_E

    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_total = st_count_E

    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_All = CA_total_E

    d1_Total_NO_m = Total_NO_m
    d1_Total_NO_f = Total_NO_f
    d1_Total_NO_All = Total_NO_All
    d1_gn_total = gn_total
    d1_bc_total = bc_total
    d1_sc_total = sc_total
    d1_st_total = st_total
    d1_CA_total_All = CA_total_All

    data_8 = {
        'SEM': ['','','',1,'','',''],
        'Course': ['','','','B.Tech P/T',' ','',''],
        'Branch': ['','','','ELECTRICAL','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_8 = pd.DataFrame(data_8)


    df_filtered_E = df[(df['SEM'] == 3) & (df['Course'].isin(['B.Tech PT'])) & (df['Branch'] == 'ELECTRICAL')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E
    Total_NO_All =  total_E

    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_total = gn_count_E

    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_total =  bc_count_E

    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_total = sc_count_E

    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_total = st_count_E

    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_All = CA_total_E

    d2_Total_NO_m = Total_NO_m
    d2_Total_NO_f = Total_NO_f
    d2_Total_NO_All = Total_NO_All
    d2_gn_total = gn_total
    d2_bc_total = bc_total
    d2_sc_total = sc_total
    d2_st_total = st_total
    d2_CA_total_All = CA_total_All

    data_9 = {
        'SEM': ['','','',3,'','',''],
        'Course': ['','','','B.Tech P/T',' ','',''],
        'Branch': ['','','','ELECTRICAL','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_9 = pd.DataFrame(data_9)

    df_filtered_E = df[(df['SEM'] == 5) & (df['Course'].isin(['B.Tech PT'])) & (df['Branch'] == 'ELECTRICAL')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E
    Total_NO_All =  total_E

    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_total = gn_count_E

    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_total =  bc_count_E

    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_total = sc_count_E

    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_total = st_count_E

    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_All = CA_total_E

    d3_Total_NO_m = Total_NO_m
    d3_Total_NO_f = Total_NO_f
    d3_Total_NO_All = Total_NO_All
    d3_gn_total = gn_total
    d3_bc_total = bc_total
    d3_sc_total = sc_total
    d3_st_total = st_total
    d3_CA_total_All = CA_total_All

    data_10 = {
        'SEM': ['','','',5,'','',''],
        'Course': ['','','','B.Tech P/T',' ','',''],
        'Branch': ['','','','ELECTRICAL','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_10 = pd.DataFrame(data_10)


    df_filtered_E = df[(df['SEM'] == 7) & (df['Course'].isin(['B.Tech PT'])) & (df['Branch'] == 'ELECTRICAL')]
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()

    Total_NO_f = f_sum_E
    total_E = m_sum_E + f_sum_E
    Total_NO_All =  total_E

    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_total = gn_count_E

    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_total =  bc_count_E

    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_total = sc_count_E

    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_total = st_count_E

    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_All = CA_total_E

    d4_Total_NO_m = Total_NO_m
    d4_Total_NO_f = Total_NO_f
    d4_Total_NO_All = Total_NO_All
    d4_gn_total = gn_total
    d4_bc_total = bc_total
    d4_sc_total = sc_total
    d4_st_total = st_total
    d4_CA_total_All = CA_total_All

    data_11 = {
        'SEM': ['','','',7,'','',''],
        'Course': ['','','','B.Tech P/T',' ','',''],
        'Branch': ['','','','ELECTRICAL','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_11 = pd.DataFrame(data_11)


    df_filtered_E = df[(df['SEM'] == 9) & (df['Course'].isin(['B.Tech PT'])) & (df['Branch'] == 'ELECTRICAL')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E
    Total_NO_All =  total_E

    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_total = gn_count_E

    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_total =  bc_count_E

    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_total = sc_count_E

    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_total = st_count_E

    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_All = CA_total_E

    d5_Total_NO_m = Total_NO_m
    d5_Total_NO_f = Total_NO_f
    d5_Total_NO_All = Total_NO_All
    d5_gn_total = gn_total
    d5_bc_total = bc_total
    d5_sc_total = sc_total
    d5_st_total = st_total
    d5_CA_total_All = CA_total_All

    GT_CA_total_All = d1_CA_total_All + d2_CA_total_All + d3_CA_total_All + d4_CA_total_All + d5_CA_total_All
    GT_st_total = d1_st_total + d2_st_total + d3_st_total + d4_st_total + d5_st_total
    GT_sc_total = d1_sc_total + d2_sc_total + d3_sc_total + d4_sc_total + d5_sc_total
    GT_bc_total = d1_bc_total + d2_bc_total + d3_bc_total + d4_bc_total + d5_bc_total
    GT_gn_total = d1_gn_total + d2_gn_total + d3_gn_total + d4_gn_total + d5_gn_total
    GT_Total_NO_All = d1_Total_NO_All + d2_Total_NO_All + d3_Total_NO_All + d4_Total_NO_All + d5_Total_NO_All
    GT_Total_NO_f = d1_Total_NO_f + d2_Total_NO_f + d3_Total_NO_f + d4_Total_NO_f + d5_Total_NO_f
    GT_Total_NO_m = d1_Total_NO_m + d2_Total_NO_m + d3_Total_NO_m + d4_Total_NO_m + d5_Total_NO_m

    Grand_Total_CA_total_All = GT_CA_total_All + Grand_Total_CA_total_All
    Grand_Total_st_total  = GT_st_total + Grand_Total_st_total
    Grand_Total_sc_total  = GT_sc_total + Grand_Total_sc_total
    Grand_Total_bc_total  = GT_bc_total + Grand_Total_bc_total
    Grand_Total_gn_total  = GT_gn_total + Grand_Total_gn_total
    Grand_Total_Total_NO_All = GT_Total_NO_All + Grand_Total_Total_NO_All
    Grand_Total_Total_NO_f = GT_Total_NO_f + Grand_Total_Total_NO_f
    Grand_Total_Total_NO_m = GT_Total_NO_m + Grand_Total_Total_NO_m

    data_12 = {
        'SEM': ['','','',9,'','',''],
        'Course': ['','','','B.Tech P/T',' ','',''],
        'Branch': ['','','','ELECTRICAL','TOTAL','GRAND TOTAL',''],
        'M': ['','','',m_sum_E,Total_NO_m,GT_Total_NO_m,''],
        'F': ['','','',f_sum_E,Total_NO_f,GT_Total_NO_f,''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,GT_Total_NO_All,''],
        'GN': ['','','',gn_count_E, gn_total,GT_gn_total,''],
        'BC': ['','','',bc_count_E, bc_total,GT_bc_total,''],
        'SC': ['','','',sc_count_E, sc_total,GT_sc_total,''],
        'ST': ['','','',st_count_E, st_total,GT_st_total,''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,GT_CA_total_All,'']
    }
    df_data_12 = pd.DataFrame(data_12)


    #----------------------------------------------------------------
    # M.Tech
    #----------------------------------------------------------------

    df_filtered_E = df[(df['SEM'] == 1) & (df['Course'].isin(['M.Tech'])) & (df['Branch'] == 'ENGINEERING SYSTEMS')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E
    Total_NO_All =  total_E

    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_total = gn_count_E

    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_total =  bc_count_E

    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_total = sc_count_E

    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_total = st_count_E

    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_All = CA_total_E

    d4_Total_NO_m = Total_NO_m
    d4_Total_NO_f = Total_NO_f
    d4_Total_NO_All = Total_NO_All
    d4_gn_total = gn_total
    d4_bc_total = bc_total
    d4_sc_total = sc_total
    d4_st_total = st_total
    d4_CA_total_All = CA_total_All

    data_13 = {
        'SEM': ['','','',1,'','',''],
        'Course': ['','','','M.Tech',' ','',''],
        'Branch': ['','','','ENGINEERING SYSTEMS','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_13 = pd.DataFrame(data_13)

    df_filtered_E = df[(df['SEM'] == 3) & (df['Course'].isin(['M.Tech'])) & (df['Branch'] == 'ENGINEERING SYSTEMS')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d5_Total_NO_m = Total_NO_m
    d5_Total_NO_f = Total_NO_f
    d5_Total_NO_All = Total_NO_All
    d5_gn_total = gn_total
    d5_bc_total = bc_total
    d5_sc_total = sc_total
    d5_st_total = st_total
    d5_CA_total_All = CA_total_All

    GT_CA_total_All = d4_CA_total_All + d5_CA_total_All
    GT_st_total = d4_st_total + d5_st_total
    GT_sc_total = d4_sc_total + d5_sc_total
    GT_bc_total = d4_bc_total + d5_bc_total
    GT_gn_total = d4_gn_total + d5_gn_total
    GT_Total_NO_All = d4_Total_NO_All + d5_Total_NO_All
    GT_Total_NO_f = d4_Total_NO_f + d5_Total_NO_f
    GT_Total_NO_m = d4_Total_NO_m + d5_Total_NO_m

    Grand_Total_CA_total_All = GT_CA_total_All + Grand_Total_CA_total_All
    Grand_Total_st_total  = GT_st_total + Grand_Total_st_total
    Grand_Total_sc_total  = GT_sc_total + Grand_Total_sc_total
    Grand_Total_bc_total  = GT_bc_total + Grand_Total_bc_total
    Grand_Total_gn_total  = GT_gn_total + Grand_Total_gn_total
    Grand_Total_Total_NO_All = GT_Total_NO_All + Grand_Total_Total_NO_All
    Grand_Total_Total_NO_f = GT_Total_NO_f + Grand_Total_Total_NO_f
    Grand_Total_Total_NO_m = GT_Total_NO_m + Grand_Total_Total_NO_m

    data_14 = {
        'SEM': ['','','',3,'','',''],
        'Course': ['','','','M.Tech',' ','',''],
        'Branch': ['','','','ENGINEERING SYSTEMS','TOTAL','GRAND TOTAL',''],
        'M': ['','','',m_sum_E,Total_NO_m,GT_Total_NO_m,''],
        'F': ['','','',f_sum_E,Total_NO_f,GT_Total_NO_f,''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,GT_Total_NO_All,''],
        'GN': ['','','',gn_count_E, gn_total,GT_gn_total,''],
        'BC': ['','','',bc_count_E, bc_total,GT_bc_total,''],
        'SC': ['','','',sc_count_E, sc_total,GT_sc_total,''],
        'ST': ['','','',st_count_E, st_total,GT_st_total,''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,GT_CA_total_All,'']
    }
    df_data_14 = pd.DataFrame(data_14)

    #----------------------------------------------------------------
    # M.Tech P/T
    #----------------------------------------------------------------

    df_filtered_E = df[(df['SEM'] == 1) & (df['Course'].isin(['M.Tech PT'])) & (df['Branch'] == 'ENGINEERING SYSTEMS')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E
    Total_NO_All =  total_E

    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_total = gn_count_E

    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_total =  bc_count_E

    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_total = sc_count_E

    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_All = CA_total_E

    d2_Total_NO_m = Total_NO_m
    d2_Total_NO_f = Total_NO_f
    d2_Total_NO_All = Total_NO_All
    d2_gn_total = gn_total
    d2_bc_total = bc_total
    d2_sc_total = sc_total
    d2_st_total = st_total
    d2_CA_total_All = CA_total_All

    data_15 = {
        'SEM': ['','','',1,'','',''],
        'Course': ['','','','M.Tech P/T',' ','',''],
        'Branch': ['','','','ENGINEERING SYSTEMS','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_15 = pd.DataFrame(data_15)

    df_filtered_E = df[(df['SEM'] == 3) & (df['Course'].isin(['M.Tech PT'])) & (df['Branch'] == 'ENGINEERING SYSTEMS')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d3_Total_NO_m = Total_NO_m
    d3_Total_NO_f = Total_NO_f
    d3_Total_NO_All = Total_NO_All
    d3_gn_total = gn_total
    d3_bc_total = bc_total
    d3_sc_total = sc_total
    d3_st_total = st_total
    d3_CA_total_All = CA_total_All

    data_16 = {
        'SEM': ['','','',3,'','',''],
        'Course': ['','','','M.Tech P/T',' ','',''],
        'Branch': ['','','','ENGINEERING SYSTEMS','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_16 = pd.DataFrame(data_16)


    df_filtered_E = df[(df['SEM'] == 5) & (df['Course'].isin(['M.Tech PT'])) & (df['Branch'] == 'ENGINEERING SYSTEMS')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d4_Total_NO_m = Total_NO_m
    d4_Total_NO_f = Total_NO_f
    d4_Total_NO_All = Total_NO_All
    d4_gn_total = gn_total
    d4_bc_total = bc_total
    d4_sc_total = sc_total
    d4_st_total = st_total
    d4_CA_total_All = CA_total_All

    data_17 = {
        'SEM': ['','','',5,'','',''],
        'Course': ['','','','M.Tech P/T',' ','',''],
        'Branch': ['','','','ENGINEERING SYSTEMS','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_17 = pd.DataFrame(data_17)


    df_filtered_E = df[(df['SEM'] == 7) & (df['Course'].isin(['M.Tech PT'])) & (df['Branch'] == 'ENGINEERING SYSTEMS')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d5_Total_NO_m = Total_NO_m
    d5_Total_NO_f = Total_NO_f
    d5_Total_NO_All = Total_NO_All
    d5_gn_total = gn_total
    d5_bc_total = bc_total
    d5_sc_total = sc_total
    d5_st_total = st_total
    d5_CA_total_All = CA_total_All

    GT_CA_total_All = d2_CA_total_All + d3_CA_total_All + d4_CA_total_All + d5_CA_total_All
    GT_st_total = d2_st_total + d3_st_total + d4_st_total + d5_st_total
    GT_sc_total = d2_sc_total + d3_sc_total + d4_sc_total + d5_sc_total
    GT_bc_total = d2_bc_total + d3_bc_total + d4_bc_total + d5_bc_total
    GT_gn_total = d2_gn_total + d3_gn_total + d4_gn_total + d5_gn_total
    GT_Total_NO_All = d2_Total_NO_All + d3_Total_NO_All + d4_Total_NO_All + d5_Total_NO_All
    GT_Total_NO_f = d2_Total_NO_f + d3_Total_NO_f + d4_Total_NO_f + d5_Total_NO_f
    GT_Total_NO_m = d2_Total_NO_m + d3_Total_NO_m + d4_Total_NO_m + d5_Total_NO_m

    Grand_Total_CA_total_All = GT_CA_total_All + Grand_Total_CA_total_All
    Grand_Total_st_total  = GT_st_total + Grand_Total_st_total
    Grand_Total_sc_total  = GT_sc_total + Grand_Total_sc_total
    Grand_Total_bc_total  = GT_bc_total + Grand_Total_bc_total
    Grand_Total_gn_total  = GT_gn_total + Grand_Total_gn_total
    Grand_Total_Total_NO_All = GT_Total_NO_All + Grand_Total_Total_NO_All
    Grand_Total_Total_NO_f = GT_Total_NO_f + Grand_Total_Total_NO_f
    Grand_Total_Total_NO_m = GT_Total_NO_m + Grand_Total_Total_NO_m

    data_18 = {
        'SEM': ['','','',7,'','',''],
        'Course': ['','','','M.Tech P/T',' ','',''],
        'Branch': ['','','','ENGINEERING SYSTEMS','TOTAL','GRAND TOTAL',''],
        'M': ['','','',m_sum_E,Total_NO_m,GT_Total_NO_m,''],
        'F': ['','','',f_sum_E,Total_NO_f,GT_Total_NO_f,''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,GT_Total_NO_All,''],
        'GN': ['','','',gn_count_E, gn_total,GT_gn_total,''],
        'BC': ['','','',bc_count_E, bc_total,GT_bc_total,''],
        'SC': ['','','',sc_count_E, sc_total,GT_sc_total,''],
        'ST': ['','','',st_count_E, st_total,GT_st_total,''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,GT_CA_total_All,'']
    }
    df_data_18 = pd.DataFrame(data_18)


    #----------------------------------------------------------------
    # M.Voc
    #----------------------------------------------------------------
    df_filtered_E = df[(df['SEM'] == 1) & (df['Course'].isin(['M.Voc'])) & (df['Branch'] == 'RENEWABLE ENERGY')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E
    Total_NO_All =  total_E

    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_total = gn_count_E

    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_total =  bc_count_E

    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_total = sc_count_E

    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d4_Total_NO_m = Total_NO_m
    d4_Total_NO_f = Total_NO_f
    d4_Total_NO_All = Total_NO_All
    d4_gn_total = gn_total
    d4_bc_total = bc_total
    d4_sc_total = sc_total
    d4_st_total = st_total
    d4_CA_total_All = CA_total_All

    data_19 = {
        'SEM': ['','','',1,'','',''],
        'Course': ['','','','M.Voc',' ','',''],
        'Branch': ['','','','RENEWABLE ENERGY','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_19 = pd.DataFrame(data_19)


    df_filtered_E = df[(df['SEM'] == 3) & (df['Course'].isin(['M.Voc'])) & (df['Branch'] == 'RENEWABLE ENERGY')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()

    Total_NO_f = f_sum_E
    total_E = m_sum_E + f_sum_E
    Total_NO_All =  total_E

    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_total = gn_count_E

    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_total =  bc_count_E

    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_total = sc_count_E

    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_total = st_count_E

    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_All = CA_total_E

    d5_Total_NO_m = Total_NO_m
    d5_Total_NO_f = Total_NO_f
    d5_Total_NO_All = Total_NO_All
    d5_gn_total = gn_total
    d5_bc_total = bc_total
    d5_sc_total = sc_total
    d5_st_total = st_total
    d5_CA_total_All = CA_total_All

    GT_CA_total_All = d4_CA_total_All + d5_CA_total_All
    GT_st_total = d4_st_total + d5_st_total
    GT_sc_total = d4_sc_total + d5_sc_total
    GT_bc_total = d4_bc_total + d5_bc_total
    GT_gn_total = d4_gn_total + d5_gn_total
    GT_Total_NO_All = d4_Total_NO_All + d5_Total_NO_All
    GT_Total_NO_f = d4_Total_NO_f + d5_Total_NO_f
    GT_Total_NO_m = d4_Total_NO_m + d5_Total_NO_m

    Grand_Total_CA_total_All = GT_CA_total_All + Grand_Total_CA_total_All
    Grand_Total_st_total  = GT_st_total + Grand_Total_st_total
    Grand_Total_sc_total  = GT_sc_total + Grand_Total_sc_total
    Grand_Total_bc_total  = GT_bc_total + Grand_Total_bc_total
    Grand_Total_gn_total  = GT_gn_total + Grand_Total_gn_total
    Grand_Total_Total_NO_All = GT_Total_NO_All + Grand_Total_Total_NO_All
    Grand_Total_Total_NO_f = GT_Total_NO_f + Grand_Total_Total_NO_f
    Grand_Total_Total_NO_m = GT_Total_NO_m + Grand_Total_Total_NO_m

    data_20 = {
        'SEM': ['','','',3,'','',''],
        'Course': ['','','','M.Voc',' ','',''],
        'Branch': ['','','','RENEWABLE ENERGY','TOTAL','GRAND TOTAL',''],
        'M': ['','','',m_sum_E,Total_NO_m,GT_Total_NO_m,''],
        'F': ['','','',f_sum_E,Total_NO_f,GT_Total_NO_f,''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,GT_Total_NO_All,''],
        'GN': ['','','',gn_count_E, gn_total,GT_gn_total,''],
        'BC': ['','','',bc_count_E, bc_total,GT_bc_total,''],
        'SC': ['','','',sc_count_E, sc_total,GT_sc_total,''],
        'ST': ['','','',st_count_E, st_total,GT_st_total,''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,GT_CA_total_All,'']
    }
    df_data_20 = pd.DataFrame(data_20)

  

    data_21 = {
        'SEM': ['','','','','','',''],
        'Course': ['','','','',' ','',''],
        'Branch': ['Grand Total of Engg','','','','','',''],
        'M': [Grand_Total_Total_NO_m ,'','','','','',''],
        'F': [Grand_Total_Total_NO_f ,'','','','','',''],
        'TOTAL(G)': [Grand_Total_Total_NO_All,'','','','','',''],
        'GN': [Grand_Total_gn_total,'','','', '','',''],
        'BC': [Grand_Total_bc_total,'','','', '','',''],
        'SC': [Grand_Total_sc_total,'','','', '','',''],
        'ST': [Grand_Total_st_total,'','','', '','',''],
        'TOTAL(C)':[Grand_Total_CA_total_All,'','','','','','']
    }

    df_data_21 = pd.DataFrame(data_21)


    blank_rows = pd.DataFrame(np.nan, index=[0, 1], columns=df_data_1.columns)
    
   
    df_combined = pd.concat([
        df_data_1, df_data_2, df_data_3, df_data_4, blank_rows,
        df_data_5, df_data_6, df_data_7, df_data_8, df_data_9, df_data_10,
        df_data_11, df_data_12, df_data_13, df_data_14, df_data_15, df_data_16,
        df_data_17, df_data_18, df_data_19, df_data_20, df_data_21
    ], ignore_index=True)

    wb = load_workbook(filename=file)
    
    sheet_name = 'Sheet3'
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    
    # Select the sheet for writing
    ws = wb[sheet_name]
    
    # Append data to the sheet
    for row in dataframe_to_rows(df_combined, index=False, header=True):
        ws.append(row)
    
    # Save changes to the workbook
    wb.save(file)

    # Return the file path or bytesIO object for downloading or further processing
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def process_excel_even(file):
    df = pd.read_excel(file, usecols="C:E,P,Q", skiprows=3, names=["SEM", "Course", "Branch", "Gender", "Category"])
    

    
    #----------------------------------------------------------------
    # B.Tech
    #----------------------------------------------------------------

    df_filtered_A = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'AGRICULTURE ENGINEERING')]
    df_filtered_C = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'CIVIL') ]
    df_filtered_E = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'ELECTRICAL')]
    df_filtered_F = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'FOOTWEAR TECHNOLOGY')]
    df_filtered_M = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'MECHANICAL')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    d1_Total_NO_m = Total_NO_m
    d1_Total_NO_f = Total_NO_f
    d1_Total_NO_All = Total_NO_All
    d1_gn_total = gn_total
    d1_bc_total = bc_total
    d1_sc_total = sc_total
    d1_st_total = st_total
    d1_CA_total_All = CA_total_All

    data_1 = {
        'SEM': [2, 2, 2, 2, 2 ,' ',''],
        'Course': ['B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', 'B.Tech',' ',''],
        'Branch': ['AGRICULTURE ENGINEERING', 'CIVIL', 'ELECTRICAL', 'FOOTWEAR TECHNOLOGY', 'MECHANICAL','TOTAL',''],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,''],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,''],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,''],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total,''],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total,''],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total,''],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total,''],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All,'']
    }


    df_data_1 = pd.DataFrame(data_1)



    df_filtered_A = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'AGRICULTURE ENGINEERING')]
    df_filtered_C = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'CIVIL') ]
    df_filtered_E = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'ELECTRICAL')]
    df_filtered_F = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'FOOTWEAR TECHNOLOGY')]
    df_filtered_M = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'MECHANICAL')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    d2_Total_NO_m = Total_NO_m
    d2_Total_NO_f = Total_NO_f
    d2_Total_NO_All = Total_NO_All
    d2_gn_total = gn_total
    d2_bc_total = bc_total
    d2_sc_total = sc_total
    d2_st_total = st_total
    d2_CA_total_All = CA_total_All

    data_2 = {
        'SEM': [4, 4, 4, 4, 4 ,' ',''],
        'Course': ['B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', 'B.Tech',' ',''],
        'Branch': ['AGRICULTURE ENGINEERING', 'CIVIL', 'ELECTRICAL', 'FOOTWEAR TECHNOLOGY', 'MECHANICAL','TOTAL',''],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,''],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,''],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,''],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total,''],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total,''],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total,''],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total,''],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All,'']
    }


    df_data_2 = pd.DataFrame(data_2)

    df_filtered_A = df[(df['SEM'] == 6) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'AGRICULTURE ENGINEERING')]
    df_filtered_C = df[(df['SEM'] == 6) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'CIVIL') ]
    df_filtered_E = df[(df['SEM'] == 6) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'ELECTRICAL')]
    df_filtered_F = df[(df['SEM'] == 6) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'FOOTWEAR TECHNOLOGY')]
    df_filtered_M = df[(df['SEM'] == 6) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'MECHANICAL')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    d3_Total_NO_m = Total_NO_m
    d3_Total_NO_f = Total_NO_f
    d3_Total_NO_All = Total_NO_All
    d3_gn_total = gn_total
    d3_bc_total = bc_total
    d3_sc_total = sc_total
    d3_st_total = st_total
    d3_CA_total_All = CA_total_All

    data_3 = {
        'SEM': [6, 6, 6, 6, 6 ,' ',''],
        'Course': ['B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', 'B.Tech',' ',''],
        'Branch': ['AGRICULTURE ENGINEERING', 'CIVIL', 'ELECTRICAL', 'FOOTWEAR TECHNOLOGY', 'MECHANICAL','TOTAL',''],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,''],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,''],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,''],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total,''],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total,''],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total,''],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total,''],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All,'']
    }

    df_data_3 = pd.DataFrame(data_3)

    df_filtered_A = df[(df['SEM'] == 8) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'AGRICULTURE ENGINEERING')]
    df_filtered_C = df[(df['SEM'] == 8) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'CIVIL') ]
    df_filtered_E = df[(df['SEM'] == 8) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'ELECTRICAL')]
    df_filtered_F = df[(df['SEM'] == 8) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'FOOTWEAR TECHNOLOGY')]
    df_filtered_M = df[(df['SEM'] == 8) & (df['Course'].isin(['B.Tech', 'B.Tech.'])) & (df['Branch'] == 'MECHANICAL')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M


    d4_Total_NO_m = Total_NO_m
    d4_Total_NO_f = Total_NO_f
    d4_Total_NO_All = Total_NO_All
    d4_gn_total = gn_total
    d4_bc_total = bc_total
    d4_sc_total = sc_total
    d4_st_total = st_total
    d4_CA_total_All = CA_total_All

    Grand_Total_CA_total_All = GT_CA_total_All = d1_CA_total_All + d2_CA_total_All + d3_CA_total_All + d4_CA_total_All
    Grand_Total_st_total  = GT_st_total = d1_st_total + d2_st_total + d3_st_total + d4_st_total
    Grand_Total_sc_total  = GT_sc_total = d1_sc_total + d2_sc_total + d3_sc_total + d4_sc_total
    Grand_Total_bc_total  = GT_bc_total = d1_bc_total + d2_bc_total + d3_bc_total + d4_bc_total
    Grand_Total_gn_total  = GT_gn_total = d1_gn_total + d2_gn_total + d3_gn_total + d4_gn_total
    Grand_Total_Total_NO_All = GT_Total_NO_All = d1_Total_NO_All + d2_Total_NO_All + d3_Total_NO_All + d4_Total_NO_All
    Grand_Total_Total_NO_f = GT_Total_NO_f = d1_Total_NO_f + d2_Total_NO_f + d3_Total_NO_f + d4_Total_NO_f
    Grand_Total_Total_NO_m = GT_Total_NO_m = d1_Total_NO_m + d2_Total_NO_m + d3_Total_NO_m + d4_Total_NO_m

    data_4 = {
        'SEM': [8, 8, 8, 8, 8 ,' ',''],
        'Course': ['B.Tech', 'B.Tech', 'B.Tech', 'B.Tech', 'B.Tech',' ',''],
        'Branch': ['AGRICULTURE ENGINEERING', 'CIVIL', 'ELECTRICAL', 'FOOTWEAR TECHNOLOGY', 'MECHANICAL','TOTAL','GRAND TOTAL'],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,GT_Total_NO_m],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,GT_Total_NO_f],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,GT_Total_NO_All],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total ,GT_gn_total],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total , GT_bc_total],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total ,GT_sc_total],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total , GT_st_total],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All , GT_CA_total_All],

    }


    df_data_4 = pd.DataFrame(data_4)


    #----------------------------------------------------------------
    # B.Voc
    #----------------------------------------------------------------
   

    df_filtered_A = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'AI & ROBOTICS')]
    df_filtered_C = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'AUTOMOBILE') ]
    df_filtered_E = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'DIGITAL MANUFACTURING')]
    df_filtered_F = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'RENEWABLE ENERGY')]
    df_filtered_M = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'WATER, SANITATION AND WASTE MANAGEMENT')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    d1_Total_NO_m = Total_NO_m
    d1_Total_NO_f = Total_NO_f
    d1_Total_NO_All = Total_NO_All
    d1_gn_total = gn_total
    d1_bc_total = bc_total
    d1_sc_total = sc_total
    d1_st_total = st_total
    d1_CA_total_All = CA_total_All

    data_5 = {
        'SEM': [2, 2, 2, 2, 2 ,' ',''],
        'Course': ['B.Voc', 'B.Voc', 'B.Voc', 'B.Voc', 'B.Voc',' ',''],
        'Branch': ['AI & ROBxOTICS', 'AUTOMOBILE', 'DIGITAL MANUFACTURING', 'RENEWABLE ENERGY', 'WATER, SANITATION AND WASTE MANAGEMENT','TOTAL',''],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,''],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,''],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,''],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total,''],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total,''],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total,''],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total,''],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All,'']
    }

    df_data_5 = pd.DataFrame(data_5)



    df_filtered_A = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'AI & ROBOTICS')]
    df_filtered_C = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'AUTOMOBILE') ]
    df_filtered_E = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'DIGITAL MANUFACTURING')]
    df_filtered_F = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'RENEWABLE ENERGY')]
    df_filtered_M = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'WATER, SANITATION AND WASTE MANAGEMENT')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    d2_Total_NO_m = Total_NO_m
    d2_Total_NO_f = Total_NO_f
    d2_Total_NO_All = Total_NO_All
    d2_gn_total = gn_total
    d2_bc_total = bc_total
    d2_sc_total = sc_total
    d2_st_total = st_total
    d2_CA_total_All = CA_total_All

    data_6 = {
        'SEM': [4, 4, 4, 4, 4 ,' ',''],
        'Course': ['B.Voc', 'B.Voc', 'B.Voc', 'B.Voc', 'B.Voc',' ',''],
        'Branch': ['AI & ROBxOTICS', 'AUTOMOBILE', 'DIGITAL MANUFACTURING', 'RENEWABLE ENERGY', 'WATER, SANITATION AND WASTE MANAGEMENT','TOTAL',''],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,''],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,''],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,''],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total,''],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total,''],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total,''],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total,''],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All,'']
    }
    df_data_6 = pd.DataFrame(data_6)

    df_filtered_A = df[(df['SEM'] == 6) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'AI & ROBOTICS')]
    df_filtered_C = df[(df['SEM'] == 6) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'AUTOMOBILE') ]
    df_filtered_E = df[(df['SEM'] == 6) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'DIGITAL MANUFACTURING')]
    df_filtered_F = df[(df['SEM'] == 6) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'RENEWABLE ENERGY')]
    df_filtered_M = df[(df['SEM'] == 6) & (df['Course'].isin(['B.Voc'])) & (df['Branch'] == 'WATER, SANITATION AND WASTE MANAGEMENT')]

    m_sum_A = df_filtered_A['Gender'].str.count('M').sum()
    m_sum_C = df_filtered_C['Gender'].str.count('M').sum()
    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()
    m_sum_F = df_filtered_F['Gender'].str.count('M').sum()
    m_sum_M = df_filtered_M['Gender'].str.count('M').sum()
    Total_NO_m = m_sum_A + m_sum_C + m_sum_E + m_sum_F + m_sum_M

    f_sum_A = df_filtered_A['Gender'].str.count('F').sum()
    f_sum_C = df_filtered_C['Gender'].str.count('F').sum()
    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()
    f_sum_F = df_filtered_F['Gender'].str.count('F').sum()
    f_sum_M = df_filtered_M['Gender'].str.count('F').sum()
    Total_NO_f = f_sum_A + f_sum_C + f_sum_E + f_sum_F + f_sum_M

    total_A = m_sum_A + f_sum_A
    total_C = m_sum_C + f_sum_C
    total_E = m_sum_E + f_sum_E
    total_F = m_sum_F + f_sum_F
    total_M = m_sum_M + f_sum_M
    Total_NO_All = total_A + total_C + total_E + total_F + total_M

    gn_count_A = df_filtered_A['Category'].str.count('GN').sum()
    gn_count_C = df_filtered_C['Category'].str.count('GN').sum()
    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_count_F = df_filtered_F['Category'].str.count('GN').sum()
    gn_count_M = df_filtered_M['Category'].str.count('GN').sum()
    gn_total = gn_count_A + gn_count_C + gn_count_E + gn_count_F + gn_count_M

    bc_count_A = df_filtered_A['Category'].str.count('BC').sum()
    bc_count_C = df_filtered_C['Category'].str.count('BC').sum()
    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_count_F = df_filtered_F['Category'].str.count('BC').sum()
    bc_count_M = df_filtered_M['Category'].str.count('BC').sum()
    bc_total = bc_count_A + bc_count_C + bc_count_E + bc_count_F + bc_count_M

    sc_count_A = df_filtered_A['Category'].str.count('SC').sum()
    sc_count_C = df_filtered_C['Category'].str.count('SC').sum()
    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_count_F = df_filtered_F['Category'].str.count('SC').sum()
    sc_count_M = df_filtered_M['Category'].str.count('SC').sum()
    sc_total = sc_count_A + sc_count_C + sc_count_E + sc_count_F + sc_count_M

    st_count_A = df_filtered_A['Category'].str.count('ST').sum()
    st_count_C = df_filtered_C['Category'].str.count('ST').sum()
    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_count_F = df_filtered_F['Category'].str.count('ST').sum()
    st_count_M = df_filtered_M['Category'].str.count('ST').sum()
    st_total = st_count_A + st_count_C + st_count_E + st_count_F + st_count_M

    CA_total_A = gn_count_A + bc_count_A + sc_count_A + st_count_A
    CA_total_C = gn_count_C + bc_count_C + sc_count_C + st_count_C
    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_F = gn_count_F + bc_count_F + sc_count_F + st_count_F
    CA_total_M = gn_count_M + bc_count_M + sc_count_M + st_count_M
    CA_total_All = CA_total_A + CA_total_C + CA_total_E + CA_total_F + CA_total_M

    d3_Total_NO_m = Total_NO_m
    d3_Total_NO_f = Total_NO_f
    d3_Total_NO_All = Total_NO_All
    d3_gn_total = gn_total
    d3_bc_total = bc_total
    d3_sc_total = sc_total
    d3_st_total = st_total
    d3_CA_total_All = CA_total_All

    GT_CA_total_All = d1_CA_total_All + d2_CA_total_All + d3_CA_total_All
    GT_st_total = d1_st_total + d2_st_total + d3_st_total
    GT_sc_total = d1_sc_total + d2_sc_total + d3_sc_total
    GT_bc_total = d1_bc_total + d2_bc_total + d3_bc_total
    GT_gn_total = d1_gn_total + d2_gn_total + d3_gn_total
    GT_Total_NO_All = d1_Total_NO_All + d2_Total_NO_All + d3_Total_NO_All
    GT_Total_NO_f = d1_Total_NO_f + d2_Total_NO_f + d3_Total_NO_f
    GT_Total_NO_m = d1_Total_NO_m + d2_Total_NO_m + d3_Total_NO_m

    Grand_Total_CA_total_All = GT_CA_total_All + Grand_Total_CA_total_All
    Grand_Total_st_total  = GT_st_total + Grand_Total_st_total
    Grand_Total_sc_total  = GT_sc_total + Grand_Total_sc_total
    Grand_Total_bc_total  = GT_bc_total + Grand_Total_bc_total
    Grand_Total_gn_total  = GT_gn_total + Grand_Total_gn_total
    Grand_Total_Total_NO_All = GT_Total_NO_All + Grand_Total_Total_NO_All
    Grand_Total_Total_NO_f = GT_Total_NO_f + Grand_Total_Total_NO_f
    Grand_Total_Total_NO_m = GT_Total_NO_m + Grand_Total_Total_NO_m


    data_7 = {
        'SEM': [6, 6, 6, 6, 6 ,' ',''],
        'Course': ['B.Voc', 'B.Voc', 'B.Voc', 'B.Voc', 'B.Voc',' ',''],
        'Branch': ['AI & ROBxOTICS', 'AUTOMOBILE', 'DIGITAL MANUFACTURING', 'RENEWABLE ENERGY', 'WATER, SANITATION AND WASTE MANAGEMENT','TOTAL','GRAND TOTAL'],
        'M': [m_sum_A, m_sum_C, m_sum_E, m_sum_F, m_sum_M,Total_NO_m,GT_Total_NO_m],
        'F': [f_sum_A, f_sum_C, f_sum_E, f_sum_F, f_sum_M,Total_NO_f,GT_Total_NO_f],
        'TOTAL(G)': [total_A, total_C, total_E, total_F, total_M,Total_NO_All,GT_Total_NO_All],
        'GN': [gn_count_A, gn_count_C, gn_count_E, gn_count_F, gn_count_M , gn_total ,GT_gn_total],
        'BC': [bc_count_A, bc_count_C, bc_count_E, bc_count_F, bc_count_M , bc_total , GT_bc_total],
        'SC': [sc_count_A, sc_count_C, sc_count_E, sc_count_F, sc_count_M , sc_total ,GT_sc_total],
        'ST': [st_count_A, st_count_C, st_count_E, st_count_F, st_count_M , st_total , GT_st_total],
        'TOTAL(C)':[CA_total_A,CA_total_C,CA_total_E,CA_total_F,CA_total_M,CA_total_All , GT_CA_total_All]
    }
    df_data_7 = pd.DataFrame(data_7)





    #----------------------------------------------------------------
    # B.Tech P/T
    #----------------------------------------------------------------




    df_filtered_E = df[(df['SEM'] == 2) & (df['Course'].isin(['B.Tech PT'])) & (df['Branch'] == 'ELECTRICAL')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d1_Total_NO_m = Total_NO_m
    d1_Total_NO_f = Total_NO_f
    d1_Total_NO_All = Total_NO_All
    d1_gn_total = gn_total
    d1_bc_total = bc_total
    d1_sc_total = sc_total
    d1_st_total = st_total
    d1_CA_total_All = CA_total_All

    data_8 = {
        'SEM': ['','','',2,'','',''],
        'Course': ['','','','B.Tech P/T',' ','',''],
        'Branch': ['','','','ELECTRICAL','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_8 = pd.DataFrame(data_8)


    df_filtered_E = df[(df['SEM'] == 4) & (df['Course'].isin(['B.Tech PT'])) & (df['Branch'] == 'ELECTRICAL')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d2_Total_NO_m = Total_NO_m
    d2_Total_NO_f = Total_NO_f
    d2_Total_NO_All = Total_NO_All
    d2_gn_total = gn_total
    d2_bc_total = bc_total
    d2_sc_total = sc_total
    d2_st_total = st_total
    d2_CA_total_All = CA_total_All

    data_9 = {
        'SEM': ['','','',4,'','',''],
        'Course': ['','','','B.Tech P/T',' ','',''],
        'Branch': ['','','','ELECTRICAL','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_9 = pd.DataFrame(data_9)

    df_filtered_E = df[(df['SEM'] == 6) & (df['Course'].isin(['B.Tech PT'])) & (df['Branch'] == 'ELECTRICAL')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d3_Total_NO_m = Total_NO_m
    d3_Total_NO_f = Total_NO_f
    d3_Total_NO_All = Total_NO_All
    d3_gn_total = gn_total
    d3_bc_total = bc_total
    d3_sc_total = sc_total
    d3_st_total = st_total
    d3_CA_total_All = CA_total_All

    data_10 = {
        'SEM': ['','','',6,'','',''],
        'Course': ['','','','B.Tech P/T',' ','',''],
        'Branch': ['','','','ELECTRICAL','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_10 = pd.DataFrame(data_10)


    df_filtered_E = df[(df['SEM'] == 8) & (df['Course'].isin(['B.Tech PT'])) & (df['Branch'] == 'ELECTRICAL')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d4_Total_NO_m = Total_NO_m
    d4_Total_NO_f = Total_NO_f
    d4_Total_NO_All = Total_NO_All
    d4_gn_total = gn_total
    d4_bc_total = bc_total
    d4_sc_total = sc_total
    d4_st_total = st_total
    d4_CA_total_All = CA_total_All

    data_11 = {
        'SEM': ['','','',8,'','',''],
        'Course': ['','','','B.Tech P/T',' ','',''],
        'Branch': ['','','','ELECTRICAL','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_11 = pd.DataFrame(data_11)


    df_filtered_E = df[(df['SEM'] == 10) & (df['Course'].isin(['B.Tech PT'])) & (df['Branch'] == 'ELECTRICAL')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d5_Total_NO_m = Total_NO_m
    d5_Total_NO_f = Total_NO_f
    d5_Total_NO_All = Total_NO_All
    d5_gn_total = gn_total
    d5_bc_total = bc_total
    d5_sc_total = sc_total
    d5_st_total = st_total
    d5_CA_total_All = CA_total_All

    GT_CA_total_All = d1_CA_total_All + d2_CA_total_All + d3_CA_total_All + d4_CA_total_All + d5_CA_total_All
    GT_st_total = d1_st_total + d2_st_total + d3_st_total + d4_st_total + d5_st_total
    GT_sc_total = d1_sc_total + d2_sc_total + d3_sc_total + d4_sc_total + d5_sc_total
    GT_bc_total = d1_bc_total + d2_bc_total + d3_bc_total + d4_bc_total + d5_bc_total
    GT_gn_total = d1_gn_total + d2_gn_total + d3_gn_total + d4_gn_total + d5_gn_total
    GT_Total_NO_All = d1_Total_NO_All + d2_Total_NO_All + d3_Total_NO_All + d4_Total_NO_All + d5_Total_NO_All
    GT_Total_NO_f = d1_Total_NO_f + d2_Total_NO_f + d3_Total_NO_f + d4_Total_NO_f + d5_Total_NO_f
    GT_Total_NO_m = d1_Total_NO_m + d2_Total_NO_m + d3_Total_NO_m + d4_Total_NO_m + d5_Total_NO_m

    Grand_Total_CA_total_All = GT_CA_total_All + Grand_Total_CA_total_All
    Grand_Total_st_total  = GT_st_total + Grand_Total_st_total
    Grand_Total_sc_total  = GT_sc_total + Grand_Total_sc_total
    Grand_Total_bc_total  = GT_bc_total + Grand_Total_bc_total
    Grand_Total_gn_total  = GT_gn_total + Grand_Total_gn_total
    Grand_Total_Total_NO_All = GT_Total_NO_All + Grand_Total_Total_NO_All
    Grand_Total_Total_NO_f = GT_Total_NO_f + Grand_Total_Total_NO_f
    Grand_Total_Total_NO_m = GT_Total_NO_m + Grand_Total_Total_NO_m

    data_12 = {
        'SEM': ['','','',10,'','',''],
        'Course': ['','','','B.Tech P/T',' ','',''],
        'Branch': ['','','','ELECTRICAL','TOTAL','GRAND TOTAL',''],
        'M': ['','','',m_sum_E,Total_NO_m,GT_Total_NO_m,''],
        'F': ['','','',f_sum_E,Total_NO_f,GT_Total_NO_f,''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,GT_Total_NO_All,''],
        'GN': ['','','',gn_count_E, gn_total,GT_gn_total,''],
        'BC': ['','','',bc_count_E, bc_total,GT_bc_total,''],
        'SC': ['','','',sc_count_E, sc_total,GT_sc_total,''],
        'ST': ['','','',st_count_E, st_total,GT_st_total,''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,GT_CA_total_All,'']
    }
    df_data_12 = pd.DataFrame(data_12)


    #----------------------------------------------------------------
    # M.Tech
    #----------------------------------------------------------------

    df_filtered_E = df[(df['SEM'] == 2) & (df['Course'].isin(['M.Tech'])) & (df['Branch'] == 'ENGINEERING SYSTEMS')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d4_Total_NO_m = Total_NO_m
    d4_Total_NO_f = Total_NO_f
    d4_Total_NO_All = Total_NO_All
    d4_gn_total = gn_total
    d4_bc_total = bc_total
    d4_sc_total = sc_total
    d4_st_total = st_total
    d4_CA_total_All = CA_total_All

    data_13 = {
        'SEM': ['','','',2,'','',''],
        'Course': ['','','','M.Tech',' ','',''],
        'Branch': ['','','','ENGINEERING SYSTEMS','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_13 = pd.DataFrame(data_13)

    df_filtered_E = df[(df['SEM'] == 4) & (df['Course'].isin(['M.Tech'])) & (df['Branch'] == 'ENGINEERING SYSTEMS')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d5_Total_NO_m = Total_NO_m
    d5_Total_NO_f = Total_NO_f
    d5_Total_NO_All = Total_NO_All
    d5_gn_total = gn_total
    d5_bc_total = bc_total
    d5_sc_total = sc_total
    d5_st_total = st_total
    d5_CA_total_All = CA_total_All

    GT_CA_total_All = d4_CA_total_All + d5_CA_total_All
    GT_st_total = d4_st_total + d5_st_total
    GT_sc_total = d4_sc_total + d5_sc_total
    GT_bc_total = d4_bc_total + d5_bc_total
    GT_gn_total = d4_gn_total + d5_gn_total
    GT_Total_NO_All = d4_Total_NO_All + d5_Total_NO_All
    GT_Total_NO_f = d4_Total_NO_f + d5_Total_NO_f
    GT_Total_NO_m = d4_Total_NO_m + d5_Total_NO_m

    Grand_Total_CA_total_All = GT_CA_total_All + Grand_Total_CA_total_All
    Grand_Total_st_total  = GT_st_total + Grand_Total_st_total
    Grand_Total_sc_total  = GT_sc_total + Grand_Total_sc_total
    Grand_Total_bc_total  = GT_bc_total + Grand_Total_bc_total
    Grand_Total_gn_total  = GT_gn_total + Grand_Total_gn_total
    Grand_Total_Total_NO_All = GT_Total_NO_All + Grand_Total_Total_NO_All
    Grand_Total_Total_NO_f = GT_Total_NO_f + Grand_Total_Total_NO_f
    Grand_Total_Total_NO_m = GT_Total_NO_m + Grand_Total_Total_NO_m

    data_14 = {
        'SEM': ['','','',4,'','',''],
        'Course': ['','','','M.Tech',' ','',''],
        'Branch': ['','','','ENGINEERING SYSTEMS','TOTAL','GRAND TOTAL',''],
        'M': ['','','',m_sum_E,Total_NO_m,GT_Total_NO_m,''],
        'F': ['','','',f_sum_E,Total_NO_f,GT_Total_NO_f,''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,GT_Total_NO_All,''],
        'GN': ['','','',gn_count_E, gn_total,GT_gn_total,''],
        'BC': ['','','',bc_count_E, bc_total,GT_bc_total,''],
        'SC': ['','','',sc_count_E, sc_total,GT_sc_total,''],
        'ST': ['','','',st_count_E, st_total,GT_st_total,''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,GT_CA_total_All,'']
    }
    df_data_14 = pd.DataFrame(data_14)

    #----------------------------------------------------------------
    # M.Tech P/T
    #----------------------------------------------------------------

    df_filtered_E = df[(df['SEM'] == 2) & (df['Course'].isin(['M.Tech PT'])) & (df['Branch'] == 'ENGINEERING SYSTEMS')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d2_Total_NO_m = Total_NO_m
    d2_Total_NO_f = Total_NO_f
    d2_Total_NO_All = Total_NO_All
    d2_gn_total = gn_total
    d2_bc_total = bc_total
    d2_sc_total = sc_total
    d2_st_total = st_total
    d2_CA_total_All = CA_total_All

    data_15 = {
        'SEM': ['','','',2,'','',''],
        'Course': ['','','','M.Tech P/T',' ','',''],
        'Branch': ['','','','ENGINEERING SYSTEMS','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_15 = pd.DataFrame(data_15)

    df_filtered_E = df[(df['SEM'] == 4) & (df['Course'].isin(['M.Tech PT'])) & (df['Branch'] == 'ENGINEERING SYSTEMS')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d3_Total_NO_m = Total_NO_m
    d3_Total_NO_f = Total_NO_f
    d3_Total_NO_All = Total_NO_All
    d3_gn_total = gn_total
    d3_bc_total = bc_total
    d3_sc_total = sc_total
    d3_st_total = st_total
    d3_CA_total_All = CA_total_All

    data_16 = {
        'SEM': ['','','',4,'','',''],
        'Course': ['','','','M.Tech P/T',' ','',''],
        'Branch': ['','','','ENGINEERING SYSTEMS','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_16 = pd.DataFrame(data_16)


    df_filtered_E = df[(df['SEM'] == 6) & (df['Course'].isin(['M.Tech PT'])) & (df['Branch'] == 'ENGINEERING SYSTEMS')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d4_Total_NO_m = Total_NO_m
    d4_Total_NO_f = Total_NO_f
    d4_Total_NO_All = Total_NO_All
    d4_gn_total = gn_total
    d4_bc_total = bc_total
    d4_sc_total = sc_total
    d4_st_total = st_total
    d4_CA_total_All = CA_total_All

    data_17 = {
        'SEM': ['','','',6,'','',''],
        'Course': ['','','','M.Tech P/T',' ','',''],
        'Branch': ['','','','ENGINEERING SYSTEMS','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_17 = pd.DataFrame(data_17)


    df_filtered_E = df[(df['SEM'] == 8) & (df['Course'].isin(['M.Tech PT'])) & (df['Branch'] == 'ENGINEERING SYSTEMS')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d5_Total_NO_m = Total_NO_m
    d5_Total_NO_f = Total_NO_f
    d5_Total_NO_All = Total_NO_All
    d5_gn_total = gn_total
    d5_bc_total = bc_total
    d5_sc_total = sc_total
    d5_st_total = st_total
    d5_CA_total_All = CA_total_All

    GT_CA_total_All = d2_CA_total_All + d3_CA_total_All + d4_CA_total_All + d5_CA_total_All
    GT_st_total = d2_st_total + d3_st_total + d4_st_total + d5_st_total
    GT_sc_total = d2_sc_total + d3_sc_total + d4_sc_total + d5_sc_total
    GT_bc_total = d2_bc_total + d3_bc_total + d4_bc_total + d5_bc_total
    GT_gn_total = d2_gn_total + d3_gn_total + d4_gn_total + d5_gn_total
    GT_Total_NO_All = d2_Total_NO_All + d3_Total_NO_All + d4_Total_NO_All + d5_Total_NO_All
    GT_Total_NO_f = d2_Total_NO_f + d3_Total_NO_f + d4_Total_NO_f + d5_Total_NO_f
    GT_Total_NO_m = d2_Total_NO_m + d3_Total_NO_m + d4_Total_NO_m + d5_Total_NO_m

    Grand_Total_CA_total_All = GT_CA_total_All + Grand_Total_CA_total_All
    Grand_Total_st_total  = GT_st_total + Grand_Total_st_total
    Grand_Total_sc_total  = GT_sc_total + Grand_Total_sc_total
    Grand_Total_bc_total  = GT_bc_total + Grand_Total_bc_total
    Grand_Total_gn_total  = GT_gn_total + Grand_Total_gn_total
    Grand_Total_Total_NO_All = GT_Total_NO_All + Grand_Total_Total_NO_All
    Grand_Total_Total_NO_f = GT_Total_NO_f + Grand_Total_Total_NO_f
    Grand_Total_Total_NO_m = GT_Total_NO_m + Grand_Total_Total_NO_m

    data_18 = {
        'SEM': ['','','',8,'','',''],
        'Course': ['','','','M.Tech P/T',' ','',''],
        'Branch': ['','','','ENGINEERING SYSTEMS','TOTAL','GRAND TOTAL',''],
        'M': ['','','',m_sum_E,Total_NO_m,GT_Total_NO_m,''],
        'F': ['','','',f_sum_E,Total_NO_f,GT_Total_NO_f,''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,GT_Total_NO_All,''],
        'GN': ['','','',gn_count_E, gn_total,GT_gn_total,''],
        'BC': ['','','',bc_count_E, bc_total,GT_bc_total,''],
        'SC': ['','','',sc_count_E, sc_total,GT_sc_total,''],
        'ST': ['','','',st_count_E, st_total,GT_st_total,''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,GT_CA_total_All,'']
    }
    df_data_18 = pd.DataFrame(data_18)


    #----------------------------------------------------------------
    # M.Voc
    #----------------------------------------------------------------
    df_filtered_E = df[(df['SEM'] == 2) & (df['Course'].isin(['M.Voc'])) & (df['Branch'] == 'RENEWABLE ENERGY')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()


    Total_NO_f = f_sum_E

    total_E = m_sum_E + f_sum_E

    Total_NO_All =  total_E


    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()

    gn_total = gn_count_E


    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()

    bc_total =  bc_count_E


    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()

    sc_total = sc_count_E


    st_count_E = df_filtered_E['Category'].str.count('ST').sum()

    st_total = st_count_E


    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E

    CA_total_All = CA_total_E

    d4_Total_NO_m = Total_NO_m
    d4_Total_NO_f = Total_NO_f
    d4_Total_NO_All = Total_NO_All
    d4_gn_total = gn_total
    d4_bc_total = bc_total
    d4_sc_total = sc_total
    d4_st_total = st_total
    d4_CA_total_All = CA_total_All

    data_19 = {
        'SEM': ['','','',2,'','',''],
        'Course': ['','','','M.Voc',' ','',''],
        'Branch': ['','','','RENEWABLE ENERGY','TOTAL','',''],
        'M': ['','','',m_sum_E,Total_NO_m,'',''],
        'F': ['','','',f_sum_E,Total_NO_f,'',''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,'',''],
        'GN': ['','','',gn_count_E, gn_total,'',''],
        'BC': ['','','',bc_count_E, bc_total,'',''],
        'SC': ['','','',sc_count_E, sc_total,'',''],
        'ST': ['','','',st_count_E, st_total,'',''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,'','']
    }
    df_data_19 = pd.DataFrame(data_19)


    df_filtered_E = df[(df['SEM'] == 4) & (df['Course'].isin(['M.Voc'])) & (df['Branch'] == 'RENEWABLE ENERGY')]

    m_sum_E = df_filtered_E['Gender'].str.count('M').sum()


    Total_NO_m = m_sum_E

    f_sum_E = df_filtered_E['Gender'].str.count('F').sum()

    Total_NO_f = f_sum_E
    total_E = m_sum_E + f_sum_E
    Total_NO_All =  total_E

    gn_count_E = df_filtered_E['Category'].str.count('GN').sum()
    gn_total = gn_count_E

    bc_count_E = df_filtered_E['Category'].str.count('BC').sum()
    bc_total =  bc_count_E

    sc_count_E = df_filtered_E['Category'].str.count('SC').sum()
    sc_total = sc_count_E

    st_count_E = df_filtered_E['Category'].str.count('ST').sum()
    st_total = st_count_E

    CA_total_E = gn_count_E + bc_count_E + sc_count_E + st_count_E
    CA_total_All = CA_total_E

    d5_Total_NO_m = Total_NO_m
    d5_Total_NO_f = Total_NO_f
    d5_Total_NO_All = Total_NO_All
    d5_gn_total = gn_total
    d5_bc_total = bc_total
    d5_sc_total = sc_total
    d5_st_total = st_total
    d5_CA_total_All = CA_total_All

    GT_CA_total_All = d4_CA_total_All + d5_CA_total_All
    GT_st_total = d4_st_total + d5_st_total
    GT_sc_total = d4_sc_total + d5_sc_total
    GT_bc_total = d4_bc_total + d5_bc_total
    GT_gn_total = d4_gn_total + d5_gn_total
    GT_Total_NO_All = d4_Total_NO_All + d5_Total_NO_All
    GT_Total_NO_f = d4_Total_NO_f + d5_Total_NO_f
    GT_Total_NO_m = d4_Total_NO_m + d5_Total_NO_m

    Grand_Total_CA_total_All = GT_CA_total_All + Grand_Total_CA_total_All
    Grand_Total_st_total  = GT_st_total + Grand_Total_st_total
    Grand_Total_sc_total  = GT_sc_total + Grand_Total_sc_total
    Grand_Total_bc_total  = GT_bc_total + Grand_Total_bc_total
    Grand_Total_gn_total  = GT_gn_total + Grand_Total_gn_total
    Grand_Total_Total_NO_All = GT_Total_NO_All + Grand_Total_Total_NO_All
    Grand_Total_Total_NO_f = GT_Total_NO_f + Grand_Total_Total_NO_f
    Grand_Total_Total_NO_m = GT_Total_NO_m + Grand_Total_Total_NO_m

    data_20 = {
        'SEM': ['','','',4,'','',''],
        'Course': ['','','','M.Voc',' ','',''],
        'Branch': ['','','','RENEWABLE ENERGY','TOTAL','GRAND TOTAL',''],
        'M': ['','','',m_sum_E,Total_NO_m,GT_Total_NO_m,''],
        'F': ['','','',f_sum_E,Total_NO_f,GT_Total_NO_f,''],
        'TOTAL(G)': ['','','',total_E,Total_NO_All,GT_Total_NO_All,''],
        'GN': ['','','',gn_count_E, gn_total,GT_gn_total,''],
        'BC': ['','','',bc_count_E, bc_total,GT_bc_total,''],
        'SC': ['','','',sc_count_E, sc_total,GT_sc_total,''],
        'ST': ['','','',st_count_E, st_total,GT_st_total,''],
        'TOTAL(C)':['','','',CA_total_E,CA_total_All,GT_CA_total_All,'']
    }
    df_data_20 = pd.DataFrame(data_20)

  

    data_21 = {
        'SEM': ['','','','','','',''],
        'Course': ['','','','',' ','',''],
        'Branch': ['Grand Total of Engg','','','','','',''],
        'M': [Grand_Total_Total_NO_m ,'','','','','',''],
        'F': [Grand_Total_Total_NO_f ,'','','','','',''],
        'TOTAL(G)': [Grand_Total_Total_NO_All,'','','','','',''],
        'GN': [Grand_Total_gn_total,'','','', '','',''],
        'BC': [Grand_Total_bc_total,'','','', '','',''],
        'SC': [Grand_Total_sc_total,'','','', '','',''],
        'ST': [Grand_Total_st_total,'','','', '','',''],
        'TOTAL(C)':[Grand_Total_CA_total_All,'','','','','','']
    }

    df_data_21 = pd.DataFrame(data_21)


    blank_rows = pd.DataFrame(np.nan, index=[0, 1], columns=df_data_1.columns)
    
   
    df_combined = pd.concat([
        df_data_1, df_data_2, df_data_3, df_data_4, blank_rows,
        df_data_5, df_data_6, df_data_7, df_data_8, df_data_9, df_data_10,
        df_data_11, df_data_12, df_data_13, df_data_14, df_data_15, df_data_16,
        df_data_17, df_data_18, df_data_19, df_data_20, df_data_21
    ], ignore_index=True)

    
    wb = load_workbook(filename=file)
    
    sheet_name = 'Sheet3'
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    
    # Select the sheet for writing
    ws = wb[sheet_name]
    
    # Append data to the sheet
    for row in dataframe_to_rows(df_combined, index=False, header=True):
        ws.append(row)
    
    # Save changes to the workbook
    wb.save(file)

    # Return the file path or bytesIO object for downloading or further processing
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
